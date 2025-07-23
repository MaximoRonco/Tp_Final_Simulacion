import tkinter as tk
from tkinter import ttk, filedialog
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import pandas as pd
from tkinter import messagebox
import math
import subprocess
import sys

def boton_moderno(master, text, command, bg, fg):
    return tk.Button(
        master, text=text, command=command,
        font=("Arial", 12, "bold"),
        bg=bg, fg=fg, activebackground=bg, activeforeground=fg,
        relief="flat", bd=0, width=18, height=2,
        highlightthickness=0,
        cursor="hand2"
    )

def van_der_pol(mu, h, n_steps, x1_0, x2_0, t0):
    t_values = [t0]
    x1_values = [x1_0]
    x2_values = [x2_0]
    tabla = []

    for i in range(n_steps + 1):
        t = t_values[-1]
        x1 = x1_values[-1]
        x2 = x2_values[-1]

        try:
            dx1_dt = x2
            dx2_dt = mu * (1 - x1**2) * x2 - x1
            x1_next = x1 + h * dx1_dt
            x2_next = x2 + h * dx2_dt
        except OverflowError:
            tabla.append((round(t,2), 'Overflow', 'Overflow', 'Overflow', 'Overflow', '-', '-'))
            break

        t_next = t + h

        if i == n_steps:
            tabla.append((round(t, 2), round(x1, 4), round(x2, 4), round(dx1_dt, 4), round(dx2_dt, 4), '-', '-'))
        else:
            tabla.append((round(t, 2), round(x1, 4), round(x2, 4), round(dx1_dt, 4), round(dx2_dt, 4), round(x1_next, 4), round(x2_next, 4)))
            t_values.append(t_next)
            x1_values.append(x1_next)
            x2_values.append(x2_next)

    return t_values, x1_values, x2_values, tabla


def graficar_resultados(t, x1, x2):
    fig, axs = plt.subplots(3, 1, figsize=(8, 10))

    axs[0].plot(t, x1, marker='o', color='blue', label='x₁(t)')
    axs[0].set_title('x₁ vs t')
    axs[0].set_xlabel('t')
    axs[0].set_ylabel('x₁')
    axs[0].grid(True)
    axs[0].legend()

    axs[1].plot(t, x2, marker='o', color='green', label="x₂(t)")
    axs[1].set_title('x₂ vs t')
    axs[1].set_xlabel('t')
    axs[1].set_ylabel('x₂')
    axs[1].grid(True)
    axs[1].legend()

    axs[2].plot(x1, x2, marker='o', color='purple', label='Fase')
    axs[2].set_title('Gráfico de fase: x₁ vs x₂')
    axs[2].set_xlabel('x₁')
    axs[2].set_ylabel('x₂')
    axs[2].grid(True)
    axs[2].legend()
    
    plt.tight_layout()
    plt.show()

def mostrar_tabla(tabla):
    for row in tree.get_children():
        tree.delete(row)
    for i, fila in enumerate(tabla):
        if i == len(tabla) - 1:
            tree.insert("", "end", values=fila, tags=('resultado',))
        else:
            tree.insert("", "end", values=fila)

def filtrar_tabla():
    try:
        desde = int(entry_desde.get()) - 1  # base 0
        hasta = int(entry_hasta.get())
    except Exception:
        desde = 0
        hasta = len(tabla_global)
    for row in tree.get_children():
        tree.delete(row)
    for i, fila in enumerate(tabla_global[desde:hasta]):
        idx_real = desde + i
        if idx_real == len(tabla_global) - 1:
            tree.insert("", "end", values=fila, tags=('resultado',))
        else:
            tree.insert("", "end", values=fila)

def exportar_tabla():
    if not tabla_global:
        messagebox.showwarning("Primero ejecutar", "⚠️ Debe ejecutar la simulación antes de exportar la tabla.")
        return
    # ... resto del código de exportación

    filetypes = [('CSV', '*.csv'), ('Excel', '*.xlsx')]
    filepath = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=filetypes)
    if not filepath:
        return

    df = pd.DataFrame(tabla_global, columns=["t", "x1", "x2", "dx1", "dx2", "x1+1", "x2+1"])
    if filepath.endswith('.csv'):
        df.to_csv(filepath, index=False)
    else:
        df.to_excel(filepath, index=False, engine='openpyxl')
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Border, Side

            wb = openpyxl.load_workbook(filepath)
            ws = wb.active

            # Encabezados en negrita y fondo
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill

            # Ajustar ancho de columnas
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2

            # Resaltar última fila (resultado)
            last_row = ws.max_row
            result_fill = PatternFill(start_color="FFE599", end_color="FFE599", fill_type="solid")
            result_font = Font(bold=True)
            for cell in ws[last_row]:
                cell.fill = result_fill
                cell.font = result_font

            # Bordes finos para toda la tabla
            thin = Side(border_style="thin", color="000000")
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            wb.save(filepath)
            wb.close()
        except ImportError:
            messagebox.showinfo("Info", "Para mejor formato en Excel, instala openpyxl: pip install openpyxl")


def validar_parametros():
    try:
        mu = float(entry_mu.get())
        h = float(entry_h.get())
        steps = int(entry_steps.get())
        x1 = float(entry_x1.get())
        x2 = float(entry_x2.get())
        t0 = float(entry_t0.get())

        if mu <= 0 or h <= 0 or steps <= 0:
            raise ValueError

        if h > 0.2:
            continuar = messagebox.askyesno(
                "Advertencia",
                f"El valor de h = {h} es grande y puede causar inestabilidad numérica.\n"
                "¿Desea continuar de todos modos con Euler?"
            )
            if not continuar:
                return False

        return True
    except:
        messagebox.showerror("Error", "⚠️ Verifique que todos los parámetros estén completos y sean válidos.")
        return False


def ejecutar_simulacion():
    global tabla_global
    if not validar_parametros():
        return
    mu = float(entry_mu.get())
    h = float(entry_h.get())
    n_steps = int(entry_steps.get())
    x1_0 = float(entry_x1.get())
    x2_0 = float(entry_x2.get())

    t0 = float(entry_t0.get())
    t, x1, x2, tabla = van_der_pol(mu, h, n_steps, x1_0, x2_0, t0)

    boton_reiniciar.config(state="normal", bg="#6c757d", fg="white")
    tabla_global = tabla
    mostrar_tabla(tabla)
    graficar_resultados(t, x1, x2)

def reiniciar_simulacion():
    global tabla_global
    tabla_global = []
    mostrar_tabla([])

    entry_mu.delete(0, tk.END)
    entry_h.delete(0, tk.END)
    entry_steps.delete(0, tk.END)
    entry_x1.delete(0, tk.END)
    entry_x2.delete(0, tk.END)
    entry_t0.delete(0, tk.END)

    entry_mu.insert(0, "2")
    entry_h.insert(0, "0.5")
    entry_steps.insert(0, "8")
    entry_x1.insert(0, "0.5")
    entry_x2.insert(0, "0")
    entry_t0.insert(0, "0")

    boton_reiniciar.config(state="disabled", bg="#cfd8dc", fg="#90a4ae")



# --- INTERFAZ MODERNA ---
root = tk.Tk()
root.title("Simulador Van der Pol - Método de Euler")
root.configure(bg="#f0f4f8")


# Estilos ttk
style = ttk.Style()
style.theme_use('clam')
style.configure("Card.TFrame", background="#e9f1f7")
style.configure("TLabel", background="#e9f1f7", font=("Arial", 11), foreground="#005f99")
style.configure("TButton", font=("Arial", 11, "bold"))
style.configure("Treeview.Heading", font=("Arial", 10, "bold"), background="#005f99")
style.configure("Treeview", font=("Arial", 10), rowheight=24)

# Botones personalizados
style.configure("Simular.TButton", font=("Arial", 11, "bold"), background="#4CAF50", foreground="white", padding=8)
style.map("Simular.TButton",
          background=[('active', '#388E3C'), ('!active', '#4CAF50')])
style.configure("Exportar.TButton", font=("Arial", 11, "bold"), background="#2196F3", foreground="white", padding=8)
style.map("Exportar.TButton",
          background=[('active', '#1565C0'), ('!active', '#2196F3')])
style.configure("Filtrar.TButton", font=("Arial", 11, "bold"), background="#FF9800", foreground="white", padding=8)
style.map("Filtrar.TButton",
          background=[('active', '#F57C00'), ('!active', '#FF9800')])
style.configure("Quitar.TButton", font=("Arial", 11, "bold"), background="#E53935", foreground="white", padding=8)
style.map("Quitar.TButton",
          background=[('active', '#B71C1C'), ('!active', '#E53935')])



# Título
titulo = tk.Label(root, text="Simulador Van der Pol - Método de Euler", font=("Arial", 18, "bold"), bg="#f0f4f8", fg="#2a4d69")
titulo.grid(row=0, column=0, columnspan=2, pady=(10, 0))

# Frame de inputs
frame = ttk.Frame(root, padding=20, style="Card.TFrame")
frame.grid(row=1, column=0, sticky="nw", padx=10, pady=10)

labels = ["mu", "h", "Iteraciones", "x1", "x2", "x (inicio de t)"]
entries = []
for i, text in enumerate(labels):
    ttk.Label(frame, text=text, font=("Arial", 10, "bold")).grid(row=i, column=0, sticky="e", pady=3)
    entry = ttk.Entry(frame, width=12)
    entry.grid(row=i, column=1, pady=3)
    entries.append(entry)

entry_mu, entry_h, entry_steps, entry_x1, entry_x2, entry_t0 = entries
entry_mu.insert(0, "2")
entry_h.insert(0, "0.5")
entry_steps.insert(0, "8")
entry_x1.insert(0, "0.5")
entry_x2.insert(0, "0")
entry_t0.insert(0, "0")  # Valor inicial de t


# Botón ejecutar
boton = boton_moderno(frame, "Ejecutar Simulación", ejecutar_simulacion, "#43a047", "white")
boton.grid(row=len(labels), columnspan=2, pady=(15, 5))


# Botón exportar

boton_exportar = boton_moderno(frame, "Exportar Tabla", exportar_tabla, "#039be5", "white")
boton_exportar.grid(row=len(labels)+1, columnspan=2, pady=5)

#Botón reiniciar
boton_reiniciar = boton_moderno(frame, "Reiniciar Simulación", reiniciar_simulacion, "#6c757d", "white")
boton_reiniciar.grid(row=len(labels)+2, columnspan=2, pady=5)
boton_reiniciar.config(state="disabled", bg="#cfd8dc", fg="#90a4ae")

# Botón volver al inicio
def volver_a_inicio():
    root.destroy()  # Cierra la ventana actual
    subprocess.run([sys.executable, "inicio.py"])  # Ejecuta inicio.py

boton_volver = boton_moderno(frame, "Volver al inicio", volver_a_inicio, "#607d8b", "white")
boton_volver.grid(row=len(labels)+3, columnspan=2, pady=5)


# Tabla con Treeview
tabla_frame = ttk.Frame(root, style="Card.TFrame")
tabla_frame.grid(row=1, column=1, padx=10, pady=10, sticky="n")

columnas = ("t", "x1", "x2", "dx1", "dx2", "x1+1", "x2+1")
tree = ttk.Treeview(tabla_frame, columns=columnas, show="headings", height=15)
for col in columnas:
    tree.heading(col, text=col)
    tree.column(col, width=90, anchor="center")
tree.pack(expand=True, fill="both")
tree.tag_configure('resultado', background='#ffe599', font=('Arial', 10, 'bold'))

# Frame para filtros
filtro_frame = ttk.Frame(root, padding=10, style="Card.TFrame")
filtro_frame.grid(row=2, column=1, sticky="w")

ttk.Label(filtro_frame, text="Desde fila:",  font=("Arial", 10, "bold")).grid(row=0, column=0, padx=2)
entry_desde = ttk.Entry(filtro_frame, width=5)
entry_desde.grid(row=0, column=1, padx=2)
ttk.Label(filtro_frame, text="Hasta fila:",  font=("Arial", 10, "bold")).grid(row=0, column=2, padx=2)
entry_hasta = ttk.Entry(filtro_frame, width=5)
entry_hasta.grid(row=0, column=3, padx=2)

boton_filtrar = boton_moderno(filtro_frame, "Filtrar tabla", filtrar_tabla, "#ff9800", "white")
boton_filtrar.grid(row=0, column=4, padx=8)



def quitar_filtros():
    mostrar_tabla(tabla_global)
    entry_desde.delete(0, tk.END)
    entry_hasta.delete(0, tk.END)
    
boton_quitar_filtros = boton_moderno(filtro_frame, "Quitar Filtros", quitar_filtros, "#e53935", "white")
boton_quitar_filtros.grid(row=0, column=5, padx=8)


# Variable global para exportar
tabla_global = []

# Centrar ventana
root.update_idletasks()
w = root.winfo_screenwidth()
h = root.winfo_screenheight()
size = tuple(int(_) for _ in root.geometry().split('+')[0].split('x'))
x = w//2 - size[0]//2
y = h//2 - size[1]//2
root.geometry("%dx%d+%d+%d" % (size + (x, y)))

root.mainloop()