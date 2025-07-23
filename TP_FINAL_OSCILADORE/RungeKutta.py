import tkinter as tk
from tkinter import ttk
import pandas as pd
import matplotlib.pyplot as plt
from tkinter import filedialog, messagebox
import subprocess
import sys


df_global = None

def van_der_pol_rk4(h, x0, x1_0, x2_0, n_iter):
    mu = 2
    results = []

    x = x0
    x1 = x1_0
    x2 = x2_0

    for i in range(n_iter):
        k1 = h * x2
        L1 = h * (mu * (1 - x1**2) * x2 - x1)

        k2 = h * (x2 + L1 / 2)
        L2 = h * (mu * (1 - (x1 + k1 / 2)**2) * (x2 + L1 / 2) - (x1 + k1 / 2))

        k3 = h * (x2 + L2 / 2)
        L3 = h * (mu * (1 - (x1 + k2 / 2)**2) * (x2 + L2 / 2) - (x1 + k2 / 2))

        k4 = h * (x2 + L3)
        L4 = h * (mu * (1 - (x1 + k3)**2) * (x2 + L3) - (x1 + k3))

        x1_next = x1 + (k1 + 2*k2 + 2*k3 + k4) / 6
        x2_next = x2 + (L1 + 2*L2 + 2*L3 + L4) / 6
        x_next = x + h

        results.append([
            i+1, round(x, 4), round(x1, 6), round(x2, 6),
            round(k1, 6), round(L1, 6), round(k2, 6), round(L2, 6),
            round(k3, 6), round(L3, 6), round(k4, 6), round(L4, 6),
            round(x_next, 4), round(x1_next, 6), round(x2_next, 6)
        ])

        x, x1, x2 = x_next, x1_next, x2_next

    columns = [
        "i", "x", "x1", "x2",
        "k1", "L1", "k2", "L2",
        "k3", "L3", "k4", "L4",
        "x(i+1)", "x1(i+1)", "x2(i+1)"
    ]
    return pd.DataFrame(results, columns=columns)

def run_simulation():
    global df_global
    if not validar_parametros_rk():
        return
    try:
        h = float(entry_h.get())
        x0 = float(entry_x.get())
        x1_0 = float(entry_x1.get())
        x2_0 = float(entry_x2.get())
        n_iter = int(entry_n.get())

        df_global = van_der_pol_rk4(h, x0, x1_0, x2_0, n_iter)
        show_table(df_global)
    except Exception as e:
        print("Error:", e)

def apply_filter():
    try:
        desde = int(entry_desde.get())
        hasta = int(entry_hasta.get())
        df_filtered = df_global.iloc[desde-1:hasta]
        show_table(df_filtered)
    except Exception as e:
        print("Error:", e)

def reset_table():
    show_table(df_global)

def show_table(df):
    for row in tree.get_children():
        tree.delete(row)

    for i, row in df.iterrows():
        tag = "last" if i == len(df_global) - 1 else ""
        tree.insert("", "end", values=list(row), tags=(tag,))

def plot_graphs():
    if df_global is not None:
        t = df_global["x"]
        x1 = df_global["x1"]
        x2 = df_global["x2"]

        plt.figure(figsize=(12, 4))

        plt.subplot(1, 3, 1)
        plt.plot(t, x1, label="x1 vs t", color="blue")
        plt.xlabel("t")
        plt.ylabel("x1")
        plt.title("x1 respecto a t")
        plt.grid()

        plt.subplot(1, 3, 2)
        plt.plot(t, x2, label="x2 vs t", color="green")
        plt.xlabel("t")
        plt.ylabel("x2")
        plt.title("x2 respecto a t")
        plt.grid()

        plt.subplot(1, 3, 3)
        plt.plot(x1, x2, label="x1 vs x2", color="red")
        plt.xlabel("x1")
        plt.ylabel("x2")
        plt.title("x1 vs x2")
        plt.grid()

        plt.tight_layout()
        plt.show()

# Interfaz
root = tk.Tk()
root.title("Método de Runge-Kutta para el Oscilador de Van der Pol")

# Estilos
style = ttk.Style()
style.configure("Treeview.Heading", font=("Arial", 10, "bold"), foreground="#005f99", background="#cceeff")
style.configure("Treeview", rowheight=22)
tree_tag = ttk.Style()
tree_tag.configure("last.Treeview", background="yellow")

def styled_label(text, row, col):
    label = tk.Label(root, text=text, font=("Arial", 10, "bold"), fg="#005f99")
    label.grid(row=row, column=col, sticky="e", padx=5, pady=2)

def styled_entry(row, col):
    entry = tk.Entry(root)
    entry.grid(row=row, column=col, padx=5, pady=2, sticky="w")
    return entry

def styled_button(text, command, row, col, color):
    btn = tk.Button(root, text=text, command=command, bg=color, fg="white",
                    font=("Arial", 10, 'bold'), relief="flat", padx=10, pady=5)
    btn.grid(row=row, column=col, sticky="ew", padx=5, pady=5)

# Título
tk.Label(root, text="Simulación del Método de Runge-Kutta (Van der Pol)", font=("Arial", 14, "bold"), fg="#003366").grid(row=0, column=0, columnspan=4, pady=10)

# Parámetros
styled_label("Paso h:", 1, 0)
entry_h = styled_entry(1, 1)

styled_label("x inicial (tiempo):", 2, 0)
entry_x = styled_entry(2, 1)

styled_label("x1 (y) inicial:", 3, 0)
entry_x1 = styled_entry(3, 1)

styled_label("x2 (y') inicial:", 4, 0)
entry_x2 = styled_entry(4, 1)

styled_label("Cantidad de iteraciones:", 5, 0)
entry_n = styled_entry(5, 1)

styled_button("Ejecutar", run_simulation, 6, 0, "#43a047")
styled_button("Graficar", plot_graphs, 6, 1, "#6f42c1")

# Tabla
cols = [
    "i", "x", "x1", "x2",
    "k1", "L1", "k2", "L2",
    "k3", "L3", "k4", "L4",
    "x(i+1)", "x1(i+1)", "x2(i+1)"
]
tree = ttk.Treeview(root, columns=cols, show="headings", height=15)
for col in cols:
    tree.heading(col, text=col)
    tree.column(col, width=80)
tree.grid(row=7, column=0, columnspan=4, padx=10, pady=10)

# Filtro debajo de la tabla
styled_label("Desde fila:", 8, 0)
entry_desde = styled_entry(8, 1)

styled_label("Hasta fila:", 8, 2)
entry_hasta = styled_entry(8, 3)

styled_button("Aplicar filtro", apply_filter, 9, 1, "#FF9800")
styled_button("Quitar filtro", reset_table, 9, 2, "#dc3545")

def validar_parametros_rk():
    try:
        h = float(entry_h.get())
        x0 = float(entry_x.get())
        x1_0 = float(entry_x1.get())
        x2_0 = float(entry_x2.get())
        n_iter = int(entry_n.get())

        if h <= 0:
            raise ValueError("El paso h debe ser mayor a 0.")
        if n_iter <= 0:
            raise ValueError("La cantidad de iteraciones debe ser un entero positivo.")

        if h > 1:
            continuar = messagebox.askyesno(
                "Advertencia",
                f"El valor de h = {h} es grande y puede causar pérdida de precisión.\n"
                "¿Desea continuar de todos modos?"
            )
            if not continuar:
                return False

        return True
    except ValueError as ve:
        messagebox.showerror("Error de entrada", f"⚠️ {ve}")
        return False
    except Exception:
        messagebox.showerror("Error", "⚠️ Verifique que todos los parámetros estén completos y sean válidos.")
        return False


def exportar_tabla():
    from tkinter import filedialog, messagebox
    if df_global is None or df_global.empty:
        messagebox.showwarning("Primero ejecutar", "⚠️ Debe ejecutar la simulación antes de exportar la tabla.")
        return
    filetypes = [('CSV', '*.csv'), ('Excel', '*.xlsx')]
    filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=filetypes)
    if not filepath or df_global is None:
        return

    df = df_global.copy()
    if filepath.endswith('.csv'):
        df.to_csv(filepath, index=False)
    else:
        df.to_excel(filepath, index=False, engine='openpyxl')
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Border, Side

            wb = openpyxl.load_workbook(filepath)
            ws = wb.active

            # Encabezados
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill

            # Ajuste de ancho
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2

            # Última fila
            last_row = ws.max_row
            result_fill = PatternFill(start_color="FFE599", end_color="FFE599", fill_type="solid")
            result_font = Font(bold=True)
            for cell in ws[last_row]:
                cell.fill = result_fill
                cell.font = result_font

            # Bordes finos
            thin = Side(border_style="thin", color="000000")
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            wb.save(filepath)
            wb.close()
        except ImportError:
            messagebox.showinfo("Info", "Para mejor formato en Excel, instala openpyxl: pip install openpyxl")
            

def volver_a_inicio():
    root.destroy()  # Cierra la ventana actual
    subprocess.run([sys.executable, "inicio.py"])  # Ejecuta inicio.py

styled_button("Volver al inicio", volver_a_inicio, 6, 3, "#607d8b")



styled_button("Exportar tabla", exportar_tabla, 6, 2, "#039be5")

root.mainloop()